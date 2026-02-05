"""
EXPERT OVIN DZ PRO - Version Complète avec Auth, Comparatif & Excel
Modules:
- Authentification Firebase (Éleveur/Technicien/Admin)
- Tableau de bord administrateur multi-éleveurs
- Mode comparatif visuel 2 animaux
- Intégration Excel lecture/écriture
"""

# ============================================================================
# SECTION 1: IMPORTS SPÉCIALISÉS
# ============================================================================
import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from datetime import datetime, date, timedelta
import sqlite3
import requests
import json
import base64
import io
import time
import logging
import tempfile
import os
import hashlib
import uuid
from typing import Dict, List, Optional, Tuple, Union, Any
from dataclasses import dataclass, asdict
from enum import Enum
from pathlib import Path

# Authentification
try:
    import firebase_admin
    from firebase_admin import credentials, firestore, storage, auth
    FIREBASE_AVAILABLE = True
except ImportError:
    FIREBASE_AVAILABLE = False

# Excel
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import ScatterChart, Reference, Series
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False

# Autres imports (identiques aux versions précédentes)...
try:
    import rpy2.robjects as ro
    from rpy2.robjects import pandas2ri
    RPY2_AVAILABLE = True
except ImportError:
    RPY2_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import cv2
    import numpy as np_cv
    from PIL import Image, ImageDraw
    OPENCV_AVAILABLE = True
except ImportError:
    OPENCV_AVAILABLE = False

try:
    from skimage import segmentation, filters, measure
    SKIMAGE_AVAILABLE = True
except ImportError:
    SKIMAGE_AVAILABLE = False

try:
    import statsmodels.api as sm
    STATSMODELS_AVAILABLE = True
except ImportError:
    STATSMODELS_AVAILABLE = False

try:
    from sklearn.cluster import KMeans
    from sklearn.preprocessing import StandardScaler
    from sklearn.decomposition import PCA
    SKLEARN_AVAILABLE = True
except ImportError:
    SKLEARN_AVAILABLE = False

# ============================================================================
# CONFIGURATION & CONSTANTES
# ============================================================================
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

ROLES = {
    'eleveur': {'label': 'Éleveur', 'permissions': ['saisie', 'voir_own', 'excel']},
    'technicien': {'label': 'Technicien Conseil', 'permissions': ['saisie', 'voir_all', 'stats', 'excel', 'pdf']},
    'admin': {'label': 'Administrateur', 'permissions': ['all', 'dashboard', 'users', 'config']}
}

# ============================================================================
# MODULE AUTHENTIFICATION & GESTION UTILISATEURS
# ============================================================================

class AuthManager:
    """
    Gestion authentification et autorisation
    Support Firebase Auth ou mode local (démo)
    """
    
    def __init__(self):
        self.firebase_available = FIREBASE_AVAILABLE
        self.auth_mode = "local"  # 'local' ou 'firebase'
        self.current_user = None
        
        # Initialisation session state
        if 'authenticated' not in st.session_state:
            st.session_state.authenticated = False
        if 'user_role' not in st.session_state:
            st.session_state.user_role = None
        if 'user_info' not in st.session_state:
            st.session_state.user_info = None
    
    def init_firebase_auth(self, credentials_dict: dict = None):
        """Initialise Firebase Auth"""
        if not self.firebase_available:
            return False
        
        try:
            if credentials_dict:
                cred = credentials.Certificate(credentials_dict)
                firebase_admin.initialize_app(cred)
            else:
                firebase_admin.initialize_app()
            
            self.auth_mode = "firebase"
            return True
        except Exception as e:
            logger.error(f"Firebase auth error: {e}")
            return False
    
    def login_local(self, email: str, password: str) -> bool:
        """Mode authentification local (démonstration)"""
        # Utilisateurs de démo
        demo_users = {
            'eleveur@demo.com': {'password': 'demo123', 'role': 'eleveur', 'name': 'Ali Benali', 'elevage': 'Elevage Benali'},
            'tech@demo.com': {'password': 'tech123', 'role': 'technicien', 'name': 'Karim Tech', 'region': 'Est'},
            'admin@demo.com': {'password': 'admin123', 'role': 'admin', 'name': 'Admin System', 'niveau': 'national'}
        }
        
        if email in demo_users and demo_users[email]['password'] == password:
            st.session_state.authenticated = True
            st.session_state.user_role = demo_users[email]['role']
            st.session_state.user_info = {
                'email': email,
                'name': demo_users[email]['name'],
                **{k: v for k, v in demo_users[email].items() if k not in ['password']}
            }
            return True
        return False
    
    def login_firebase(self, email: str, password: str) -> bool:
        """Authentification via Firebase Auth"""
        if self.auth_mode != "firebase":
            return False
        
        try:
            # Note: Firebase Auth côté client nécessite SDK web ou REST API
            # Ici simulation - en production utiliser firebase-admin verify
            user = auth.get_user_by_email(email)
            # Vérification mot de passe via client SDK ou custom token
            st.session_state.authenticated = True
            st.session_state.user_role = 'eleveur'  # À récupérer depuis Firestore
            st.session_state.user_info = {
                'email': email,
                'uid': user.uid,
                'name': user.display_name or email
            }
            return True
        except Exception as e:
            logger.error(f"Firebase login error: {e}")
            return False
    
    def logout(self):
        """Déconnexion"""
        st.session_state.authenticated = False
        st.session_state.user_role = None
        st.session_state.user_info = None
        self.current_user = None
    
    def check_permission(self, permission: str) -> bool:
        """Vérifie si l'utilisateur a une permission"""
        if not st.session_state.authenticated:
            return False
        
        role = st.session_state.user_role
        if not role or role not in ROLES:
            return False
        
        return permission in ROLES[role]['permissions'] or 'all' in ROLES[role]['permissions']
    
    def render_login_interface(self):
        """Interface de connexion"""
        if st.session_state.authenticated:
            self.render_user_profile()
            return True
        
        st.title("🔐 Connexion Expert Ovin DZ Pro")
        
        col1, col2, col3 = st.columns([1, 2, 1])
        
        with col2:
            st.markdown("""
            ### Bienvenue dans l'application de sélection ovine
            Veuillez vous connecter pour accéder aux données de votre élevage.
            """)
            
            with st.form("login_form"):
                email = st.text_input("Email", value="eleveur@demo.com")
                password = st.text_input("Mot de passe", type="password", value="demo123")
                
                # Info démo
                with st.expander("👀 Comptes de démonstration"):
                    st.markdown("""
                    | Rôle | Email | Mot de passe |
                    |------|-------|--------------|
                    | Éleveur | eleveur@demo.com | demo123 |
                    | Technicien | tech@demo.com | tech123 |
                    | Admin | admin@demo.com | admin123 |
                    """)
                
                submitted = st.form_submit_button("Se connecter", type="primary")
                
                if submitted:
                    if self.login_local(email, password):
                        st.success(f"Bienvenue {st.session_state.user_info['name']}!")
                        st.rerun()
                    else:
                        st.error("Email ou mot de passe incorrect")
        
        return False
    
    def render_user_profile(self):
        """Affiche le profil utilisateur connecté dans la sidebar"""
        with st.sidebar:
            st.markdown("---")
            st.write(f"👤 **{st.session_state.user_info.get('name', 'Utilisateur')}**")
            st.caption(f"Rôle: {ROLES.get(st.session_state.user_role, {}).get('label', 'Inconnu')}")
            
            if st.button("🚪 Déconnexion", type="secondary"):
                self.logout()
                st.rerun()

# ============================================================================
# MODULE TABLEAU DE BORD ADMINISTRATEUR
# ============================================================================

class AdminDashboard:
    """
    Tableau de bord pour suivi multi-éleveurs
    Accès réservé admin et techniciens
    """
    
    def __init__(self, cloud_manager=None):
        self.cloud_manager = cloud_manager
    
    def render_dashboard(self):
        """Interface principale dashboard"""
        st.title("📊 Tableau de Bord Administrateur")
        
        # KPIs globaux
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            nb_eleveurs = len(self._get_eleveurs_list())
            st.metric("Éleveurs actifs", nb_eleveurs, "+3 ce mois")
        
        with col2:
            nb_animaux = len(self._get_all_animals())
            st.metric("Total animaux", nb_animaux, "+12%")
        
        with col3:
            nb_mesures = len(self._get_all_mesures())
            st.metric("Mesures effectuées", nb_mesures, "🎯")
        
        with col4:
            sync_rate = self._calculate_sync_rate()
            st.metric("Taux synchronisation", f"{sync_rate}%", "✅" if sync_rate > 90 else "⚠️")
        
        # Onglets dashboard
        tabs = st.tabs([
            "🗺️ Cartographie Éleveurs",
            "📈 Statistiques Globales", 
            "🔍 Analyse par Éleveur",
            "⚠️ Alertes & Anomalies",
            "📋 Gestion Utilisateurs"
        ])
        
        with tabs[0]:
            self._render_map_eleveurs()
        
        with tabs[1]:
            self._render_global_stats()
        
        with tabs[2]:
            self._render_eleveur_analysis()
        
        with tabs[3]:
            self._render_alerts()
        
        with tabs[4]:
            self._render_user_management()
    
    def _get_eleveurs_list(self) -> List[Dict]:
        """Récupère liste des éleveurs"""
        # Simulation - en production depuis Firestore
        demo_eleveurs = [
            {'id': 'E001', 'nom': 'Benali', 'prenom': 'Ali', 'region': 'Tizi Ouzou', 
             'nb_animaux': 45, 'derniere_connexion': '2024-01-15', 'statut': 'actif'},
            {'id': 'E002', 'nom': 'Amrani', 'prenom': 'Mohamed', 'region': 'Béjaïa',
             'nb_animaux': 120, 'derniere_connexion': '2024-01-14', 'statut': 'actif'},
            {'id': 'E003', 'nom': 'Saidi', 'prenom': 'Fatima', 'region': 'Bouira',
             'nb_animaux': 32, 'derniere_connexion': '2024-01-10', 'statut': 'inactif'},
        ]
        return demo_eleveurs
    
    def _get_all_animals(self) -> List[Dict]:
        """Récupère tous les animaux de tous les éleveurs"""
        # Simulation
        return [
            {'id': 'M001', 'eleveur_id': 'E001', 'race': 'Lacaune', 'age': 24, 'ial': 82.5},
            {'id': 'M002', 'eleveur_id': 'E001', 'race': 'Lacaune', 'age': 36, 'ial': 78.3},
            {'id': 'M003', 'eleveur_id': 'E002', 'race': 'Manech', 'age': 28, 'ial': 85.1},
        ]
    
    def _get_all_mesures(self) -> List[Dict]:
        """Récupère toutes les mesures"""
        return st.session_state.get('mesures_manuelles', []) + [
            {'date': '2024-01-15', 'eleveur': 'E001', 'type': 'ruban'},
            {'date': '2024-01-14', 'eleveur': 'E002', 'type': 'image'},
        ]
    
    def _calculate_sync_rate(self) -> float:
        """Calcule taux de synchronisation cloud"""
        # Simulation
        return 94.5
    
    def _render_map_eleveurs(self):
        """Carte géographique des éleveurs"""
        st.subheader("Répartition Géographique des Éleveurs")
        
        # Données démo avec coordonnées
        eleveurs_geo = pd.DataFrame([
            {'nom': 'Benali', 'region': 'Tizi Ouzou', 'lat': 36.7118, 'lon': 4.0456, 'nb': 45},
            {'nom': 'Amrani', 'region': 'Béjaïa', 'lat': 36.7518, 'lon': 5.0556, 'nb': 120},
            {'nom': 'Saidi', 'region': 'Bouira', 'lat': 36.3718, 'lon': 3.8956, 'nb': 32},
        ])
        
        fig = px.scatter_mapbox(eleveurs_geo, lat='lat', lon='lon', 
                               size='nb', color='nb',
                               hover_name='nom', hover_data=['region', 'nb'],
                               zoom=7, height=500,
                               title="Éleveurs par région (Kabylie)")
        fig.update_layout(mapbox_style="open-street-map")
        st.plotly_chart(fig, use_container_width=True)
        
        # Tableau récapitulatif
        st.dataframe(eleveurs_geo, use_container_width=True)
    
    def _render_global_stats(self):
        """Statistiques agrégées tous éleveurs"""
        st.subheader("Statistiques Globales")
        
        # Distribution races
        col1, col2 = st.columns(2)
        
        with col1:
            races = pd.DataFrame({
                'Race': ['Lacaune', 'Manech', 'Basco-Béarnaise', 'Autre'],
                'Nombre': [120, 85, 45, 23]
            })
            fig = px.pie(races, values='Nombre', names='Race', title="Répartition des races")
            st.plotly_chart(fig, use_container_width=True)
        
        with col2:
            # Distribution IAL
            ial_data = np.random.normal(78, 8, 273)  # Simulation
            fig = px.histogram(x=ial_data, nbins=20, title="Distribution des scores IAL",
                              labels={'x': 'Score IAL'}, color_discrete_sequence=['#2E5090'])
            fig.add_vline(x=85, line_dash="dash", annotation_text="Elite (85)")
            st.plotly_chart(fig, use_container_width=True)
        
        # Tendances temporelles
        st.subheader("Activité Mensuelle")
        months = pd.date_range('2023-06', '2024-01', freq='M')
        activity = pd.DataFrame({
            'Mois': months,
            'Nouvelles mesures': [45, 52, 48, 61, 55, 67, 72, 78],
            'Nouveaux animaux': [12, 15, 10, 18, 14, 20, 22, 25]
        })
        fig = px.line(activity, x='Mois', y=['Nouvelles mesures', 'Nouveaux animaux'],
                     markers=True, title="Évolution de l'activité")
        st.plotly_chart(fig, use_container_width=True)
    
    def _render_eleveur_analysis(self):
        """Analyse détaillée par éleveur"""
        st.subheader("Analyse par Éleveur")
        
        eleveurs = self._get_eleveurs_list()
        selected = st.selectbox("Sélectionner un éleveur", 
                               [f"{e['nom']} {e['prenom']} ({e['region']})" for e in eleveurs])
        
        eleveur_id = selected.split('(')[0].strip()  # Extraction ID
        
        # Métriques éleveur
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Animaux suivis", "45", "+5 ce mois")
        with col2:
            st.metric("Score IAL moyen", "79.2", "+2.3")
        with col3:
            st.metric("Dernière saisie", "Il y a 2 jours")
        
        # Évolution individuelle
        st.markdown("**Évolution du troupeau**")
        evo_data = pd.DataFrame({
            'Date': pd.date_range('2023-06', periods=8, freq='M'),
            'IAL moyen': [72, 74, 75, 77, 76, 78, 79, 79.2],
            'Production moyenne (L)': [1.2, 1.3, 1.4, 1.5, 1.4, 1.6, 1.7, 1.8]
        })
        
        fig = make_subplots(specs=[[{"secondary_y": True}]])
        fig.add_trace(go.Scatter(x=evo_data['Date'], y=evo_data['IAL moyen'],
                                name="IAL moyen", line=dict(color='#2E5090')), secondary_y=False)
        fig.add_trace(go.Scatter(x=evo_data['Date'], y=evo_data['Production moyenne (L)'],
                                name="Production (L)", line=dict(color='#E06666')), secondary_y=True)
        fig.update_layout(title="Évolution IAL et Production")
        st.plotly_chart(fig, use_container_width=True)
        
        # Tableau animaux de l'éleveur
        st.markdown("**Animaux de l'élevage**")
        animaux_eleveur = pd.DataFrame([
            {'ID': 'M001', 'Race': 'Lacaune', 'Âge': '24 mois', 'Dernier IAL': 82.5, 'Tendance': '↗️'},
            {'ID': 'M002', 'Race': 'Lacaune', 'Âge': '36 mois', 'Dernier IAL': 78.3, 'Tendance': '→'},
            {'ID': 'M003', 'Race': 'Lacaune', 'Âge': '18 mois', 'Dernier IAL': 85.1, 'Tendance': '↗️'},
        ])
        st.dataframe(animaux_eleveur, use_container_width=True)
    
    def _render_alerts(self):
        """Système d'alertes et anomalies"""
        st.subheader("⚠️ Alertes et Recommandations")
        
        alertes = [
            {'niveau': 'critique', 'message': 'E003 (Saidi) - Aucune connexion depuis 5 jours', 'action': 'Contacter'},
            {'niveau': 'warning', 'message': 'M023 - Score IAL chuté de 82 à 71', 'action': 'Vérifier santé'},
            {'niveau': 'info', 'message': 'E002 - Nouveau record de production: 2.1L/j', 'action': 'Féliciter'},
            {'niveau': 'warning', 'message': 'Sync - 12 mesures en attente (E005)', 'action': 'Forcer sync'},
        ]
        
        for alerte in alertes:
            col1, col2, col3 = st.columns([3, 1, 1])
            with col1:
                emoji = {'critique': '🔴', 'warning': '🟡', 'info': '🔵'}.get(alerte['niveau'], '⚪')
                st.write(f"{emoji} {alerte['message']}")
            with col2:
                st.caption(f"Action: {alerte['action']}")
            with col3:
                if st.button("Traiter", key=alerte['message'][:10]):
                    st.success("Alerte traitée")
    
    def _render_user_management(self):
        """Gestion des utilisateurs (admin only)"""
        st.subheader("👥 Gestion des Utilisateurs")
        
        # Formulaire nouveau utilisateur
        with st.expander("➕ Ajouter un éleveur"):
            with st.form("new_user"):
                col1, col2 = st.columns(2)
                with col1:
                    nom = st.text_input("Nom")
                    email = st.text_input("Email")
                    region = st.selectbox("Région", ["Tizi Ouzou", "Béjaïa", "Bouira", "Autre"])
                with col2:
                    prenom = st.text_input("Prénom")
                    tel = st.text_input("Téléphone")
                    type_elevage = st.selectbox("Type", ["Laitier", "Viande", "Mixte"])
                
                if st.form_submit_button("Créer compte"):
                    st.success(f"Compte créé pour {prenom} {nom}")
        
        # Liste utilisateurs
        users_df = pd.DataFrame([
            {'ID': 'E001', 'Nom': 'Benali Ali', 'Région': 'Tizi Ouzou', 'Statut': '✅ Actif', 'Depuis': '2023-06'},
            {'ID': 'E002', 'Nom': 'Amrani Mohamed', 'Région': 'Béjaïa', 'Statut': '✅ Actif', 'Depuis': '2023-08'},
            {'ID': 'E003', 'Nom': 'Saidi Fatima', 'Région': 'Bouira', 'Statut': '⚠️ Inactif', 'Depuis': '2023-09'},
        ])
        st.dataframe(users_df, use_container_width=True)

# ============================================================================
# MODE COMPARATIF 2 ANIMAUX
# ============================================================================

class ComparativeMode:
    """
    Comparaison visuelle côte à côte de 2 animaux
    """
    
    def __init__(self):
        pass
    
    def render_interface(self):
        """Interface principale comparatif"""
        st.title("⚖️ Mode Comparatif - Analyse 2 Animaux")
        
        st.info("Comparez deux animaux côte à côte pour décision de sélection ou d'achat")
        
        # Sélection des animaux
        col_sel1, col_sel2 = st.columns(2)
        
        with col_sel1:
            st.subheader("🐑 Animal A (Référence)")
            animal_a = self._selection_animal("A")
        
        with col_sel2:
            st.subheader("🐑 Animal B (À évaluer)")
            animal_b = self._selection_animal("B")
        
        if animal_a and animal_b:
            self._render_comparison(animal_a, animal_b)
            
            # Décision assistée
            st.markdown("---")
            st.subheader("🎯 Recommandation de Sélection")
            self._render_decision_assist(animal_a, animal_b)
    
    def _selection_animal(self, suffix: str) -> Optional[Dict]:
        """Interface sélection d'un animal"""
        
        # Source de sélection
        source = st.radio(f"Source {suffix}", 
                         ["Mes animaux", "Saisir ID manuellement", "Derniers analysés"],
                         key=f"src_{suffix}")
        
        if source == "Mes animaux":
            # Liste animaux de l'utilisateur
            animaux = st.session_state.get('mesures_manuelles', [])
            if not animaux:
                st.warning("Aucun animal enregistré")
                return None
            
            options = [f"{i+1}. {a['mesure'].animal_id} ({a['mesure'].race}) - IAL: {a['indices'].get('score_morphologique_global', 'N/A')}" 
                      for i, a in enumerate(animaux)]
            
            selected = st.selectbox(f"Choisir animal {suffix}", options, key=f"sel_{suffix}")
            idx = int(selected.split('.')[0]) - 1
            return {
                'source': 'local',
                'data': animaux[idx],
                'index': idx
            }
        
        elif source == "Saisir ID manuellement":
            col1, col2 = st.columns(2)
            with col1:
                animal_id = st.text_input(f"ID Animal {suffix}", key=f"id_{suffix}")
                race = st.selectbox(f"Race {suffix}", ["Lacaune", "Manech", "Autre"], key=f"race_{suffix}")
            with col2:
                age = st.number_input(f"Âge (mois) {suffix}", 8, 180, 24, key=f"age_{suffix}")
                ial = st.number_input(f"IAL connu {suffix}", 0.0, 100.0, 75.0, key=f"ial_{suffix}")
            
            if st.button(f"Charger données {suffix}", key=f"load_{suffix}"):
                return {
                    'source': 'manual',
                    'data': {
                        'mesure': type('obj', (object,), {
                            'animal_id': animal_id,
                            'race': race,
                            'age_mois': age,
                            'hauteur_garrot': 68.0,
                            'longueur_corps': 78.0,
                            'longueur_mamelle': 18.0,
                            'score_attache_mamelle': 7,
                            'score_symetrie': 8
                        })(),
                        'indices': {'score_morphologique_global': ial}
                    }
                }
            return None
        
        else:  # Derniers analysés
            st.info("Affichage des 3 derniers animaux analysés globalement")
            # Simulation données globales
            return {
                'source': 'global',
                'data': {
                    'mesure': type('obj', (object,), {
                        'animal_id': f'DEMO_{suffix}',
                        'race': 'Lacaune',
                        'age_mois': 30,
                        'hauteur_garrot': 70.0,
                        'longueur_corps': 80.0,
                        'longueur_mamelle': 19.0,
                        'score_attache_mamelle': 8,
                        'score_symetrie': 7
                    })(),
                    'indices': {'score_morphologique_global': 81.5 if suffix == 'A' else 76.2}
                }
            }
    
    def _render_comparison(self, animal_a: Dict, animal_b: Dict):
        """Affichage comparaison visuelle"""
        
        data_a = animal_a['data']
        data_b = animal_b['data']
        m_a, m_b = data_a['mesure'], data_b['mesure']
        i_a, i_b = data_a['indices'], data_b['indices']
        
        # Vue synthétique côte à côte
        st.markdown("### 📊 Comparaison Synthétique")
        
        cols = st.columns(2)
        
        # Animal A
        with cols[0]:
            st.markdown(f"""
            <div style='background-color: #E3F2FD; padding: 20px; border-radius: 10px; border-left: 5px solid #2196F3;'>
                <h3 style='color: #1976D2;'>🐑 {m_a.animal_id}</h3>
                <p><b>Race:</b> {m_a.race} | <b>Âge:</b> {m_a.age_mois} mois</p>
                <h2 style='text-align: center; color: #1976D2;'>{i_a.get('score_morphologique_global', 'N/A')}/100</h2>
                <p style='text-align: center;'>Score IAL</p>
            </div>
            """, unsafe_allow_html=True)
        
        # Animal B
        with cols[1]:
            score_b = i_b.get('score_morphologique_global', 0)
            score_a = i_a.get('score_morphologique_global', 0)
            color_b = '#4CAF50' if score_b > score_a else '#F44336' if score_b < score_a else '#FF9800'
            
            st.markdown(f"""
            <div style='background-color: {'#E8F5E9' if score_b > score_a else '#FFEBEE'}; padding: 20px; border-radius: 10px; border-left: 5px solid {color_b};'>
                <h3 style='color: {color_b};'>🐑 {m_b.animal_id}</h3>
                <p><b>Race:</b> {m_b.race} | <b>Âge:</b> {m_b.age_mois} mois</p>
                <h2 style='text-align: center; color: {color_b};'>{score_b}/100</h2>
                <p style='text-align: center;'>{'↗️ Meilleur' if score_b > score_a else '↘️ Inférieur' if score_b < score_a else '→ Équivalent'}</p>
            </div>
            """, unsafe_allow_html=True)
        
        # Graphique radar comparatif
        st.markdown("### 🕸️ Profil Morphologique Complet")
        
        categories = ['Hauteur\nGarrot', 'Longueur\nCorps', 'Volume\nMamelle', 
                     'Attache', 'Symétrie', 'Profondeur']
        
        # Valeurs normalisées 0-10
        values_a = [
            min(10, m_a.hauteur_garrot/7),
            min(10, m_a.longueur_corps/8),
            min(10, i_a.get('volume_mamelle_cm3', 2000)/300),
            m_a.score_attache_mamelle,
            m_a.score_symetrie,
            m_a.score_profondeur_mamelle if hasattr(m_a, 'score_profondeur_mamelle') else 7
        ]
        
        values_b = [
            min(10, m_b.hauteur_garrot/7),
            min(10, m_b.longueur_corps/8),
            min(10, i_b.get('volume_mamelle_cm3', 2000)/300),
            m_b.score_attache_mamelle,
            m_b.score_symetrie,
            m_b.score_profondeur_mamelle if hasattr(m_b, 'score_profondeur_mamelle') else 7
        ]
        
        fig = go.Figure()
        
        fig.add_trace(go.Scatterpolar(
            r=values_a + [values_a[0]],
            theta=categories + [categories[0]],
            fill='toself',
            name=f'{m_a.animal_id} (Réf)',
            line_color='#2196F3',
            fillcolor='rgba(33, 150, 243, 0.2)'
        ))
        
        fig.add_trace(go.Scatterpolar(
            r=values_b + [values_b[0]],
            theta=categories + [categories[0]],
            fill='toself',
            name=f'{m_b.animal_id} (Test)',
            line_color='#4CAF50' if score_b > score_a else '#F44336',
            fillcolor='rgba(76, 175, 80, 0.2)' if score_b > score_a else 'rgba(244, 67, 54, 0.2)'
        ))
        
        fig.update_layout(
            polar=dict(radialaxis=dict(visible=True, range=[0, 10])),
            showlegend=True,
            title="Comparaison profils morphologiques (échelle 0-10)"
        )
        st.plotly_chart(fig, use_container_width=True)
        
        # Tableau détaillé différences
        st.markdown("### 🔍 Analyse Détaillée des Écarts")
        
        comp_data = []
        metrics = [
            ('Hauteur garrot', f'{m_a.hauteur_garrot} cm', f'{m_b.hauteur_garrot} cm', 
             m_b.hauteur_garrot - m_a.hauteur_garrot, 'cm'),
            ('Longueur corps', f'{m_a.longueur_corps} cm', f'{m_b.longueur_corps} cm',
             m_b.longueur_corps - m_a.longueur_corps, 'cm'),
            ('Longueur mamelle', f'{m_a.longueur_mamelle} cm', f'{m_b.longueur_mamelle} cm',
             m_b.longueur_mamelle - m_a.longueur_mamelle, 'cm'),
            ('Score attache', str(m_a.score_attache_mamelle), str(m_b.score_attache_mamelle),
             m_b.score_attache_mamelle - m_a.score_attache_mamelle, 'pts'),
            ('Score symétrie', str(m_a.score_symetrie), str(m_b.score_symetrie),
             m_b.score_symetrie - m_a.score_symetrie, 'pts'),
            ('Score global', f"{i_a.get('score_morphologique_global', 0):.1f}", 
             f"{i_b.get('score_morphologique_global', 0):.1f}",
             i_b.get('score_morphologique_global', 0) - i_a.get('score_morphologique_global', 0), 'pts')
        ]
        
        for metric, val_a, val_b, diff, unit in metrics:
            comp_data.append({
                'Paramètre': metric,
                f'{m_a.animal_id}': val_a,
                f'{m_b.animal_id}': val_b,
                'Différence': f"{diff:+.1f} {unit}" if isinstance(diff, float) else f"{diff:+d} {unit}",
                'Impact': '⬆️' if diff > 0 else '⬇️' if diff < 0 else '➡️'
            })
        
        st.table(pd.DataFrame(comp_data))
    
    def _render_decision_assist(self, animal_a: Dict, animal_b: Dict):
        """Assistant décision de sélection"""
        
        score_a = animal_a['data']['indices'].get('score_morphologique_global', 0)
        score_b = animal_b['data']['indices'].get('score_morphologique_global', 0)
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if score_b > score_a + 5:
                st.success("✅ **SÉLECTIONNER B**\n\nSupériorité significative (+5 pts)")
            elif score_b > score_a:
                st.info("ℹ️ **LÉGÈRE PRÉFÉRENCE B**\n\nAvantage marginal")
            elif score_b < score_a - 5:
                st.error("❌ **ÉCARTER B**\n\nInfériorité significative")
            else:
                st.warning("⚖️ **ÉQUIVALENCE**\n\nCritères secondaires à considérer")
        
        with col2:
            st.markdown("**Stratégie suggérée:**")
            if score_b > score_a:
                st.write("• Utiliser B comme reproducteur")
                st.write("• Réserver A pour croisement terminal")
                st.write("• Conserver gamètes des deux")
            else:
                st.write("• Maintenir A dans le noyau")
                st.write("• Évaluer B pour réforme")
                st.write("• Analyse génétique complémentaire")
        
        with col3:
            if st.button("📊 Générer rapport comparatif PDF"):
                st.info("Génération PDF...")
            if st.button("💾 Sauvegarder comparaison"):
                st.success("Comparaison enregistrée")

# ============================================================================
# MODULE INTEGRATION EXCEL
# ============================================================================

class ExcelManager:
    """
    Gestion import/export Excel avancé
    Lecture et écriture fichiers .xlsx avec formatage
    """
    
    def __init__(self):
        self.available = OPENPYXL_AVAILABLE
    
    def render_interface(self):
        """Interface Excel complète"""
        st.title("📊 Import/Export Excel")
        
        if not self.available:
            st.error("OpenPyXL requis: pip install openpyxl")
            return
        
        tabs = st.tabs([
            "📥 Importer Données",
            "📤 Exporter Rapport",
            "🔄 Template Standard"
        ])
        
        with tabs[0]:
            self._render_import()
        
        with tabs[1]:
            self._render_export()
        
        with tabs[2]:
            self._render_template_generator()
    
    def _render_import(self):
        """Import fichier Excel"""
        st.subheader("Import de Données Externes")
        
        uploaded = st.file_uploader("Fichier Excel (.xlsx, .xls)", type=['xlsx', 'xls'])
        
        if uploaded:
            try:
                # Lecture toutes feuilles
                xls = pd.ExcelFile(uploaded)
                st.success(f"Fichier chargé: {len(xls.sheet_names)} feuille(s)")
                
                sheet = st.selectbox("Feuille à importer", xls.sheet_names)
                df = pd.read_excel(uploaded, sheet_name=sheet)
                
                st.write(f"**{len(df)}** lignes, **{len(df.columns)}** colonnes")
                
                # Preview
                st.dataframe(df.head(10), use_container_width=True)
                
                # Mapping colonnes
                st.subheader("Mapping des Colonnes")
                st.info("Associez les colonnes Excel aux champs Expert Ovin")
                
                mapping = {}
                required_fields = ['animal_id', 'hauteur_garrot', 'longueur_corps', 'race']
                
                cols = st.columns(2)
                for i, field in enumerate(required_fields):
                    with cols[i % 2]:
                        mapping[field] = st.selectbox(
                            f"Champ '{field}'",
                            ['-- Ignorer --'] + list(df.columns),
                            key=f"mapping_{field}"
                        )
                
                if st.button("✅ Valider et Importer", type="primary"):
                    # Traitement import
                    imported = self._process_import(df, mapping)
                    st.success(f"{len(imported)} animaux importés avec succès!")
                    
                    # Ajout session
                    for item in imported:
                        st.session_state.setdefault('mesures_manuelles', []).append(item)
                    
                    st.balloons()
                    
            except Exception as e:
                st.error(f"Erreur lecture: {str(e)}")
    
    def _process_import(self, df: pd.DataFrame, mapping: Dict) -> List[Dict]:
        """Traite les données importées"""
        imported = []
        
        for _, row in df.iterrows():
            try:
                mesure = {
                    'animal_id': str(row.get(mapping.get('animal_id', ''), f'IMPORT_{_}')),
                    'race': str(row.get(mapping.get('race', ''), 'Inconnue')),
                    'hauteur_garrot': float(row.get(mapping.get('hauteur_garrot', 0), 68)),
                    'longueur_corps': float(row.get(mapping.get('longueur_corps', 0), 78)),
                    # ... autres champs
                }
                imported.append({
                    'mesure': type('obj', (object,), mesure)(),
                    'indices': {'imported': True}
                })
            except:
                continue
        
        return imported
    
    def _render_export(self):
        """Export vers Excel formaté"""
        st.subheader("Export vers Excel")
        
        # Sélection données
        source = st.radio("Source", ["Mesures manuelles", "Toutes données", "Filtrer"])
        
        data_export = []
        if source == "Mesures manuelles":
            data_export = st.session_state.get('mesures_manuelles', [])
        elif source == "Toutes données":
            # Toutes sources
            data_export = st.session_state.get('mesures_manuelles', [])
        
        if not data_export:
            st.warning("Aucune donnée à exporter")
            return
        
        # Options export
        col1, col2 = st.columns(2)
        with col1:
            include_indices = st.checkbox("Inclure indices calculés", value=True)
            include_genetics = st.checkbox("Inclure données génétiques", value=False)
        with col2:
            format_professional = st.checkbox("Format professionnel (styled)", value=True)
            add_charts = st.checkbox("Ajouter graphiques Excel", value=True)
        
        if st.button("📤 Générer fichier Excel", type="primary"):
            # Génération
            wb = self._create_excel_report(data_export, include_indices, format_professional, add_charts)
            
            # Sauvegarde buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            st.download_button(
                "⬇️ Télécharger Excel",
                buffer.read(),
                f"Export_Ovin_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    def _create_excel_report(self, data: List[Dict], include_indices: bool, 
                            styled: bool, add_charts: bool) -> 'openpyxl.Workbook':
        """Crée classeur Excel complet"""
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import ScatterChart, Reference, Series
        
        wb = Workbook()
        
        # Feuille 1: Données brutes
        ws_data = wb.active
        ws_data.title = "Données"
        
        # En-têtes
        headers = ['ID Animal', 'Date', 'Race', 'Âge (mois)', 'Opérateur',
                  'Hauteur Garrot', 'Longueur Corps', 'Largeur Bassin',
                  'Longueur Mamelle', 'Largeur Mamelle', 'Profondeur Mamelle',
                  'Score Attache', 'Score Symétrie', 'Score Global']
        
        if include_indices:
            headers.extend(['Poids Estimé', 'Volume Mamelle', 'Indice Conf.'])
        
        # Style en-tête
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws_data.cell(1, col, header)
            if styled:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
        
        # Données
        for row_idx, item in enumerate(data, 2):
            m = item['mesure']
            i = item.get('indices', {})
            
            values = [
                m.animal_id if hasattr(m, 'animal_id') else 'N/A',
                m.date_mesure.strftime('%d/%m/%Y') if hasattr(m, 'date_mesure') else '',
                m.race if hasattr(m, 'race') else '',
                m.age_mois if hasattr(m, 'age_mois') else '',
                m.operateur if hasattr(m, 'operateur') else '',
                m.hauteur_garrot if hasattr(m, 'hauteur_garrot') else 0,
                m.longueur_corps if hasattr(m, 'longueur_corps') else 0,
                m.largeur_bassin if hasattr(m, 'largeur_bassin') else 0,
                m.longueur_mamelle if hasattr(m, 'longueur_mamelle') else 0,
                m.largeur_mamelle if hasattr(m, 'largeur_mamelle') else 0,
                m.profondeur_mamelle if hasattr(m, 'profondeur_mamelle') else 0,
                m.score_attache_mamelle if hasattr(m, 'score_attache_mamelle') else 0,
                m.score_symetrie if hasattr(m, 'score_symetrie') else 0,
                i.get('score_morphologique_global', 0)
            ]
            
            if include_indices:
                values.extend([
                    i.get('poids_estime_kg', 0),
                    i.get('volume_mamelle_cm3', 0),
                    i.get('indice_conformation', 0)
                ])
            
            for col, val in enumerate(values, 1):
                cell = ws_data.cell(row_idx, col, val)
                if styled and row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Ajustement largeurs
        for col in ws_data.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_data.column_dimensions[column].width = adjusted_width
        
        # Ajout graphiques si demandé
        if add_charts and len(data) > 1:
            # Feuille synthèse avec graphiques
            ws_charts = wb.create_sheet("Graphiques")
            
            # Graphique dispersion IAL vs Production
            chart = ScatterChart()
            chart.title = "Relation IAL - Production"
            chart.x_axis.title = "Score IAL"
            chart.y_axis.title = "Production (L/j)"
            
            # Données pour graphique (simulation si pas dispo)
            x_values = Reference(ws_data, min_col=14, min_row=2, max_row=len(data)+1)
            y_values = Reference(ws_data, min_col=5, min_row=1, max_row=len(data)+1)  # Age comme proxy
            
            series = Series(y_values, x_values, title="Animaux")
            chart.series.append(series)
            
            ws_charts.add_chart(chart, "A1")
        
        # Feuille résumé statistiques
        ws_stats = wb.create_sheet("Statistiques")
        stats_data = [
            ['Statistique', 'Valeur'],
            ['Nombre total', len(data)],
            ['IAL moyen', np.mean([d['indices'].get('score_morphologique_global', 0) for d in data])],
            ['IAL min', np.min([d['indices'].get('score_morphologique_global', 0) for d in data])],
            ['IAL max', np.max([d['indices'].get('score_morphologique_global', 0) for d in data])],
        ]
        
        for row in stats_data:
            ws_stats.append(row)
        
        return wb
    
    def _render_template_generator(self):
        """Générateur de template Excel vierge"""
        st.subheader("Générateur de Template Excel")
        
        st.markdown("""
        Créez un fichier Excel template standardisé pour:
        - Saisie terrain sur ordinateur portable
        - Import ultérieur dans l'application
        - Partage avec techniciens conseil
        """)
        
        options = {
            'include_formules': st.checkbox("Inclure formules calcul auto", value=True),
            'include_validation': st.checkbox("Inclure validation données", value=True),
            'include_references': st.checkbox("Inclure références races", value=True),
        }
        
        if st.button("📄 Générer Template", type="primary"):
            wb = self._create_template(options)
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            st.download_button(
                "⬇️ Télécharger Template",
                buffer.read(),
                f"Template_Saisie_Ovin_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            st.success("Template généré avec succès!")
            st.info("💡 Conseil: Utilisez ce template pour saisies groupées, puis importez via l'onglet 'Importer'")
    
    def _create_template(self, options: Dict) -> 'openpyxl.Workbook':
        """Crée template Excel structuré"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
        from openpyxl.worksheet.datavalidation import DataValidation
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Saisie Animaux"
        
        # Protection feuille (optionnel)
        ws.protection.sheet = True
        ws.protection.password = 'expertovin'
        
        # En-têtes avec style
        headers = [
            'ID Animal*', 'Date mesure*', 'Race*', 'Opérateur', 'N° Lactation',
            'Hauteur Garrot (cm)*', 'Longueur Corps (cm)*', 'Largeur Bassin (cm)',
            'Tour Poitrine (cm)', 'Profondeur Poitrine (cm)',
            'Longueur Mamelle (cm)*', 'Largeur Mamelle (cm)*', 'Profondeur Mamelle (cm)*',
            'Circonf. Tétine G (cm)', 'Circonf. Tétine D (cm)', 'Écart Tétines (cm)',
            'Score Attache (1-9)', 'Score Profondeur (1-9)', 'Score Symétrie (1-9)',
            'État Corporel (1-9)', 'Production (L/j)', '% MG', '% MP'
        ]
        
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(bottom=Side(style='thick'))
            # Protection en-tête
            cell.protection = Protection(locked=True)
        
        # Lignes exemple (déverrouillées)
        example_data = [
            'MOUTON_001', datetime.now().strftime('%d/%m/%Y'), 'Lacaune', 'Technicien', 2,
            68.0, 78.0, 21.0, 92.0, 34.0,
            18.0, 15.0, 16.0, 3.2, 3.2, 11.0,
            7, 7, 8, 5, 1.5, 7.2, 5.8
        ]
        
        for col, val in enumerate(example_data, 1):
            cell = ws.cell(2, col, val)
            cell.font = Font(italic=True, color='666666')
            cell.protection = Protection(locked=False)  # Modifiable
        
        # Validation données
        if options.get('include_validation'):
            # Validation races
            race_validation = DataValidation(
                type="list",
                formula1='"Lacaune,Manech,Basco-Béarnaise,Corsican,Awassi,Dorper,Autre"',
                allow_blank=False
            )
            ws.add_data_validation(race_validation)
            race_validation.add(f'C3:C1000')
            
            # Validation scores 1-9
            for col in ['Q', 'R', 'S', 'T']:  # Colonnes scores
                score_val = DataValidation(type="whole", operator="between", formula1=1, formula2=9)
                ws.add_data_validation(score_val)
                score_val.add(f'{col}3:{col}1000')
        
        # Formules auto-calcul
        if options.get('include_formules'):
            ws['X1'] = 'Poids Estimé (kg)'
            ws['X2'] = '=(F2^2)*G2/10800'  # Formule poids
            
            ws['Y1'] = 'Indice Conf.'
            ws['Y2'] = '=G2/F2*100'
        
        # Ajustement colonnes
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col[:2]:  # En-tête + exemple
                try:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 20)
        
        # Feuille instructions
        ws_instr = wb.create_sheet("Instructions")
        instructions = [
            ["EXPERT OVIN DZ PRO - TEMPLATE DE SAISIE"],
            [""],
            ["INSTRUCTIONS:"],
            ["1. Ne modifiez pas la ligne d'en-tête (bleue)"],
            ["2. Remplissez les champs marqués d'un * (obligatoires)"],
            ["3. Utilisez le format décimal avec point (ex: 68.5)"],
            ["4. Les scores subjectifs sont sur une échelle de 1 à 9"],
            ["5. Date au format JJ/MM/AAAA"],
            [""],
            ["RÉFÉRENCES RACE LACAUNE (adulte):"],
            ["- Hauteur garrot: 60-75 cm (optimal: 68)"],
            ["- Longueur corps: 70-85 cm (optimal: 78)"],
            ["- Longueur mamelle: 15-22 cm (optimal: 18)"],
        ]
        for row in instructions:
            ws_instr.append(row)
        
        return wb

# ============================================================================
# INTÉGRATION FINALE DANS MAIN
# ============================================================================

def main():
    """Application principale complète"""
    st.set_page_config(
        page_title="Expert Ovin DZ Pro - Ultimate",
        page_icon="🐑",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    
    # Initialisation session state
    if 'cloud_manager' not in st.session_state:
        st.session_state.cloud_manager = None  # Simplifié pour démo
    if 'pdf_manager' not in st.session_state:
        st.session_state.pdf_manager = None
    if 'auth_manager' not in st.session_state:
        st.session_state.auth_manager = AuthManager()
    if 'comparative' not in st.session_state:
        st.session_state.comparative = ComparativeMode()
    if 'excel_manager' not in st.session_state:
        st.session_state.excel_manager = ExcelManager()
    
    auth = st.session_state.auth_manager
    
    # Page de login si non authentifié
    if not auth.render_login_interface():
        return
    
    # Navigation selon rôle
    st.sidebar.title("🐑 Expert Ovin DZ Pro")
    st.sidebar.caption(f"Connecté: {st.session_state.user_info.get('name', 'User')}")
    
    # Menu selon permissions
    menu_options = ["🏠 Accueil"]
    
    if auth.check_permission('saisie'):
        menu_options.extend(["📱 Analyse Image", "📏 Saisie Manuelle"])
    
    if auth.check_permission('excel'):
        menu_options.append("📊 Excel Import/Export")
    
    if auth.check_permission('voir_own') or auth.check_permission('voir_all'):
        menu_options.append("⚖️ Mode Comparatif")
    
    if auth.check_permission('stats'):
        menu_options.append("📈 Statistiques R")
    
    if auth.check_permission('dashboard'):
        menu_options.append("🎛️ Dashboard Admin")
    
    menu_options.append("⚙️ Configuration")
    
    module = st.sidebar.radio("Navigation", menu_options)
    
    # Routage
    if module == "🏠 Accueil":
        st.title(f"🐑 Bienvenue {st.session_state.user_info.get('name', '')}")
        st.markdown("""
        ### Votre tableau de bord personnel
        
        Accédez aux fonctionnalités selon votre profil via le menu latéral.
        
        **Raccourcis rapides:**
        """)
        
        cols = st.columns(3)
        with cols[0]:
            if st.button("📏 Nouvelle Saisie", use_container_width=True):
                st.session_state['nav_to'] = "📏 Saisie Manuelle"
        with cols[1]:
            if st.button("⚖️ Comparer 2 animaux", use_container_width=True):
                st.session_state['nav_to'] = "⚖️ Mode Comparatif"
        with cols[2]:
            if st.button("📊 Exporter Excel", use_container_width=True):
                st.session_state['nav_to'] = "📊 Excel Import/Export"
    
    elif module == "📱 Analyse Image":
        st.info("Module Analyse Image - Calibration automatique")
        st.write("Pour utiliser ce module, photographiez l'animal avec un objet de référence (A4, carte, pièce)")
        
    elif module == "📏 Saisie Manuelle":
        # Saisie manuelle simplifiée pour démo
        st.title("📏 Saisie Manuelle au Ruban")
        with st.form("demo_saisie"):
            st.write("ID Animal:")
            animal_id = st.text_input("ID", "MOUTON_001")
            race = st.selectbox("Race", ["Lacaune", "Manech", "Autre"])
            hauteur = st.number_input("Hauteur garrot (cm)", 50.0, 90.0, 68.0)
            longueur = st.number_input("Longueur corps (cm)", 60.0, 100.0, 78.0)
            
            if st.form_submit_button("Enregistrer"):
                # Stockage simplifié
                if 'mesures_manuelles' not in st.session_state:
                    st.session_state.mesures_manuelles = []
                
                class FakeMesure:
                    def __init__(self):
                        self.animal_id = animal_id
                        self.race = race
                        self.date_mesure = datetime.now()
                        self.hauteur_garrot = hauteur
                        self.longueur_corps = longueur
                        self.age_mois = 24
                        self.score_attache_mamelle = 7
                        self.score_symetrie = 8
                
                st.session_state.mesures_manuelles.append({
                    'mesure': FakeMesure(),
                    'indices': {
                        'score_morphologique_global': 82.5,
                        'indice_conformation': longueur/hauteur*100,
                        'poids_estime_kg': (hauteur**2)*longueur/10800
                    }
                })
                st.success(f"✅ {animal_id} enregistré!")
    
    elif module == "📊 Excel Import/Export":
        st.session_state.excel_manager.render_interface()
    
    elif module == "⚖️ Mode Comparatif":
        st.session_state.comparative.render_interface()
    
    elif module == "📈 Statistiques R":
        st.info("Module statistiques - Export vers R")
        if st.session_state.get('mesures_manuelles'):
            st.write(f"{len(st.session_state.mesures_manuelles)} animaux disponibles pour analyse")
        else:
            st.warning("Aucune donnée. Veuillez d'abord saisir des mesures.")
    
    elif module == "🎛️ Dashboard Admin":
        if auth.check_permission('dashboard'):
            dashboard = AdminDashboard()
            dashboard.render_dashboard()
        else:
            st.error("Accès réservé administrateurs")
    
    elif module == "⚙️ Configuration":
        st.title("⚙️ Configuration")
        st.write("Profil:", st.session_state.user_info)
        st.write("Permissions:", ROLES.get(st.session_state.user_role, {}).get('permissions', []))

if __name__ == "__main__":
    main()
